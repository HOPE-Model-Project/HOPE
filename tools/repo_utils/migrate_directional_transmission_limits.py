#!/usr/bin/env python3
"""Migrate HOPE transmission inputs to strict forward/reverse capacity columns.

The migration is intentionally breaking:

* ``Capacity (MW)`` is removed from linedata, linedata_candidate, branchdata,
  and saved transmission-build ``line.csv`` outputs.
* ``Forward Capacity (MW)`` and ``Reverse Capacity (MW)`` are required.
* Symmetric legacy ratings are copied into both directional columns.
* ICARUS tables that preserve Z1->Z2 and Z2->Z1 source ratings use those values.

CSV files and relevant sheets in XLSX workbooks are migrated in place. Run with
one or more repository/case roots.
"""

from __future__ import annotations

import argparse
import csv
import math
from copy import copy
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


FORWARD = "Forward Capacity (MW)"
REVERSE = "Reverse Capacity (MW)"
LEGACY = "Capacity (MW)"
ICARUS_FORWARD = "ICARUS_Cap_Z1_to_Z2_MW"
ICARUS_REVERSE = "ICARUS_Cap_Z2_to_Z1_MW"
INPUT_TABLE_NAMES = {"linedata", "linedata_candidate", "branchdata"}
CSV_TABLE_NAMES = INPUT_TABLE_NAMES | {"line"}


def transmission_csvs(root: Path) -> Iterable[Path]:
    for path in root.rglob("*.csv"):
        if path.stem.lower() in CSV_TABLE_NAMES:
            yield path


def input_workbooks(root: Path) -> Iterable[Path]:
    for path in root.rglob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        if "input" in path.stem.lower():
            yield path


def parse_rating(value: object, *, location: str) -> float:
    if value is None or (isinstance(value, str) and not value.strip()):
        raise ValueError(f"{location}: directional transmission rating is blank")
    try:
        rating = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"{location}: directional transmission rating is not numeric: {value!r}"
        ) from exc
    if not math.isfinite(rating) or rating < 0.0:
        raise ValueError(
            f"{location}: directional transmission rating must be finite and non-negative; "
            f"found {rating}"
        )
    return rating


def validate_csv_ratings(path: Path, rows: list[list[str]]) -> None:
    header = rows[0]
    forward_idx = header.index(FORWARD)
    reverse_idx = header.index(REVERSE)
    for row_number, row in enumerate(rows[1:], start=2):
        if len(row) != len(header):
            raise ValueError(
                f"{path}:{row_number}: expected {len(header)} fields, found {len(row)}"
            )
        parse_rating(row[forward_idx], location=f"{path}:{row_number} [{FORWARD}]")
        parse_rating(row[reverse_idx], location=f"{path}:{row_number} [{REVERSE}]")


def migrate_csv(path: Path) -> bool:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    if not rows:
        raise ValueError(f"{path}: transmission CSV is empty and has no header")

    header = rows[0]
    has_forward = FORWARD in header
    has_reverse = REVERSE in header
    has_legacy = LEGACY in header

    if has_forward or has_reverse:
        if not (has_forward and has_reverse):
            raise ValueError(f"{path}: only one directional capacity column is present")
        if has_legacy:
            raise ValueError(f"{path}: mixes legacy and directional capacity columns")
        validate_csv_ratings(path, rows)
        return False
    if not has_legacy:
        raise ValueError(f"{path}: no recognized transmission-capacity column")

    legacy_idx = header.index(LEGACY)
    icarus_forward_idx = header.index(ICARUS_FORWARD) if ICARUS_FORWARD in header else None
    icarus_reverse_idx = header.index(ICARUS_REVERSE) if ICARUS_REVERSE in header else None
    if (icarus_forward_idx is None) != (icarus_reverse_idx is None):
        raise ValueError(f"{path}: incomplete preserved ICARUS directional source columns")

    new_rows: list[list[str]] = []
    new_header = header[:legacy_idx] + [FORWARD, REVERSE] + header[legacy_idx + 1 :]
    new_rows.append(new_header)
    for row_number, row in enumerate(rows[1:], start=2):
        if len(row) != len(header):
            raise ValueError(
                f"{path}:{row_number}: expected {len(header)} fields, found {len(row)}"
            )
        if icarus_forward_idx is None:
            forward_value = row[legacy_idx]
            reverse_value = row[legacy_idx]
        else:
            forward_value = row[icarus_forward_idx]
            reverse_value = row[icarus_reverse_idx]
        new_rows.append(
            row[:legacy_idx]
            + [forward_value, reverse_value]
            + row[legacy_idx + 1 :]
        )

    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerows(new_rows)
    validate_csv_ratings(path, new_rows)
    return True


def copy_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def migrate_worksheet(path: Path, worksheet) -> bool:
    header_values = [cell.value for cell in worksheet[1]]
    normalized = [str(value).strip() if value is not None else "" for value in header_values]
    has_forward = FORWARD in normalized
    has_reverse = REVERSE in normalized
    has_legacy = LEGACY in normalized

    if has_forward or has_reverse:
        if not (has_forward and has_reverse):
            raise ValueError(
                f"{path} [{worksheet.title}]: only one directional capacity column is present"
            )
        if has_legacy:
            raise ValueError(
                f"{path} [{worksheet.title}]: mixes legacy and directional columns"
            )
        forward_col = normalized.index(FORWARD) + 1
        reverse_col = normalized.index(REVERSE) + 1
        for row in range(2, worksheet.max_row + 1):
            parse_rating(
                worksheet.cell(row, forward_col).value,
                location=f"{path} [{worksheet.title}!{worksheet.cell(row, forward_col).coordinate}]",
            )
            parse_rating(
                worksheet.cell(row, reverse_col).value,
                location=f"{path} [{worksheet.title}!{worksheet.cell(row, reverse_col).coordinate}]",
            )
        return False
    if not has_legacy:
        raise ValueError(
            f"{path} [{worksheet.title}]: no recognized transmission-capacity column"
        )

    legacy_col = normalized.index(LEGACY) + 1
    worksheet.cell(1, legacy_col).value = FORWARD
    worksheet.insert_cols(legacy_col + 1, 1)
    worksheet.cell(1, legacy_col + 1).value = REVERSE
    copy_cell_style(worksheet.cell(1, legacy_col), worksheet.cell(1, legacy_col + 1))
    for row in range(2, worksheet.max_row + 1):
        source = worksheet.cell(row, legacy_col)
        target = worksheet.cell(row, legacy_col + 1)
        target.value = source.value
        copy_cell_style(source, target)
    return True


def migrate_workbook(path: Path) -> tuple[bool, list[str]]:
    workbook = load_workbook(path)
    changed_sheets: list[str] = []
    for sheet_name in workbook.sheetnames:
        if sheet_name.strip().lower() not in INPUT_TABLE_NAMES:
            continue
        worksheet = workbook[sheet_name]
        if migrate_worksheet(path, worksheet):
            changed_sheets.append(sheet_name)
    if changed_sheets:
        workbook.save(path)
    return bool(changed_sheets), changed_sheets


def audit_root(root: Path) -> tuple[int, int]:
    csv_changed = 0
    workbook_changed = 0
    for path in sorted(transmission_csvs(root)):
        if migrate_csv(path):
            csv_changed += 1
            print(f"CSV  {path}")
    for path in sorted(input_workbooks(root)):
        changed, sheets = migrate_workbook(path)
        if changed:
            workbook_changed += 1
            print(f"XLSX {path} [{', '.join(sheets)}]")
    return csv_changed, workbook_changed


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("roots", nargs="+", type=Path)
    args = parser.parse_args()

    total_csv = 0
    total_xlsx = 0
    for root in args.roots:
        resolved = root.resolve()
        if not resolved.is_dir():
            raise NotADirectoryError(resolved)
        csv_changed, xlsx_changed = audit_root(resolved)
        total_csv += csv_changed
        total_xlsx += xlsx_changed
    print(f"Migrated {total_csv} CSV files and {total_xlsx} XLSX workbooks.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
